import pandas as pd
import numpy as np
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from customer.models import Customer, Bag, Container


class Command(BaseCommand):
    help = 'Generate a random Excel file with right-to-left layout'

    def handle(self, *args, **kwargs):
        # Generate random data for the first sheet
        num_rows = Customer.objects.all().count()
        customers = Customer.objects.all()
        data = []
        data2 = []
        i = 1
        for customer in customers:
            formatted_pk = str(customer.pk).zfill(5)
            formatted_date_joined = customer.date_joined.strftime('%y%m%d')
            bags = Bag.objects.filter(customer=customer)
            bags_count = sum(bag.number for bag in bags)  
            bags_weight = sum(bag.weight for bag in bags)  
            container = Container.objects.get(customer=customer)

            data.append({
                'الرقم': i,
                'رقم التسجيل': f'{formatted_pk}-{formatted_date_joined}',
                'تاريخ التسجيل': customer.date_joined.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S'),
                'رقم الهاتف 1': customer.phone,
                'الولاية': customer.state.state,
                'المنطقة': customer.zone.zone,
                'عدد الاكياس': bags_count,
                'الوزن ( كغ )': bags_weight,
                'عدد الاوعية': container.number,
                'السعة ( لتر )': container.capacity,
                'عدد الايام منذ الجني': customer.days_gone,
                'ملاحضات': '',
                'العامل ': customer.user.username if customer.user is not None else '',
                'تم تعديله': 'نعم' if customer.modified else 'لا',
                'تاريخ التعديل': customer.modification_date.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if customer.modified and customer.modification_date else '',
                'العامل': customer.modification_user.username if customer.modified else '',
                'ملغى ': 'نعم' if not customer.is_active else 'لا',
                'تاريخ الإلغاء ': customer.cancel_date.replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if not customer.is_active else '',
                'سبب الإلغاء ': customer.cancel_reason if not customer.is_active else '',
                'ملاحظات إضافية ': '',
                'العامل  ': customer.cancel_user.username if not customer.is_active else '',
            })
            j=1
            bag_data = {
                'الرقم': j,  
            }
            for index, bag in enumerate(bags, start=1):
                bag_data[f'الاكياس {index}'] = bag.number
            for index, bag in enumerate(bags, start=1):
                bag_data[f'الوزن  {index}'] = bag.weight

            data2.append(bag_data)

            i += 1

        # Create a DataFrame for the first sheet
        df = pd.DataFrame(data)

        # Create a new workbook and add a worksheet for the first sheet
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = 'ov'

        # Set the sheet's properties for right-to-left
        sheet1.sheet_view.rightToLeft = True

        # Write the DataFrame to the first worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet1.cell(row=r_idx, column=c_idx, value=value)

        # Adjust column widths for the first sheet
        # sheet1.column_dimensions['B'].width = 13
        # sheet1.column_dimensions['C'].width = 18
        # sheet1.column_dimensions['D'].width = 11
        # sheet1.column_dimensions['O'].width = 18
        # sheet1.column_dimensions['Q'].width = 18
        # sheet1.column_dimensions['R'].width = 12
        # sheet1.column_dimensions['S'].width = 13
        for col in sheet1.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet1.column_dimensions[column].width = adjusted_width

        # Generate random data for the second sheet
        

        # Create a DataFrame for the random data
        df2 = pd.DataFrame(data2)

        # Add a new worksheet for the second sheet
        sheet2 = workbook.create_sheet(title='dtov')

        # Set the sheet's properties for right-to-left
        sheet2.sheet_view.rightToLeft = True

        # Write the random DataFrame to the second worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df2, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet2.cell(row=r_idx, column=c_idx, value=value)

        # Adjust column widths for the second sheet
        for col in sheet2.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet2.column_dimensions[column].width = adjusted_width

        # Save to Excel
        file_path = 'C:/Users/OMEN 16/Desktop/olivePress/backend/xlsx_files/D1.xlsx'  # Define your file path
        workbook.save(file_path)

        self.stdout.write(self.style.SUCCESS('D1 Excel file generated successfully!'))
